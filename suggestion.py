# import openpyxl module
import openpyxl
import xlsxwriter
from openpyxl.utils.cell import get_column_letter

path = "excel.xlsx"
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active

cell_list = []
exact_list = []

def suggest(a, b):
    i = 2
    x = 2
    # finding movie
    data = sheet_obj.cell(row = 1, column = i)
    while data != a:
        i+=1
        data = sheet_obj.cell(row = 1, column = i)

    number = sheet_obj.cell(row = x, column = i)
    # finding all reviews for movie
    while i < 20000: # change for the amount of rows you have
        if number != None:
            cell_list.append(str(i))
            if number == b:
                exact_list.append(str(i))
        x+=1
        number = sheet_obj.cell(row = x, column = i)

    ratings = {'Movie': [], 'Rating': []}

    if len(exact_list) != 0:
        # make suggestion based off of that list
        for i in exact_list:
            loop_val = i
            loop = 0
            letter = get_column_letter(loop)
            movie_name = sheet_obj.cell(row = 1, column = letter)
            rating = sheet_obj.cell(row = i, column = letter)
            while loop < 15000:
                if(rating!=None):
                    ratings[movie_name] = rating
                    loop += 1
                else:
                    loop += 1
              
                letter = get_column_letter(loop)
                movie_name = sheet_obj.cell(row = 1, column = letter)
                rating = sheet_obj.cell(row = i, column = letter)


    elif len(cell_list) != 0:
        #find closest review
        print(False)
    else:
        print("no reviews")




    return True

print(suggest('Home Alone', 10))

