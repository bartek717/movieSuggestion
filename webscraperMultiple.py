#To get the url, and scrape the html page  
import requests
from bs4 import BeautifulSoup
#To save the reviews in a dataframe 
import pandas as pd
import selenium
import time
from selenium import webdriver
import smtplib
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import load_workbook
import xlsxwriter
import openpyxl
from openpyxl.utils.cell import get_column_letter



# webdriver stuff
browser = "C:\webdrivers/chromedriver"

options = Options()
options.add_argument("--incognito")
options.add_argument("--disable-blink-features=AutomationControlled")
options.add_argument("window-size=1200x600")

options.add_argument("start-maximized")
driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)

# excel shit
workbook = xlsxwriter.Workbook('excel.xlsx')
worksheet = workbook.add_worksheet()

init_url = 'https://www.metacritic.com/browse/movies/score/metascore/all/filtered?sort=desc&page=0'

user_agent = {'User-agent': 'Mozilla/5.0'}
response = requests.get(init_url, headers = user_agent)
soup = BeautifulSoup(response.text, 'html.parser')

review_dict = {'name':[], 'rating':[]}
movie_arr = []
user_arr = []

#look for review_content tags since every user review has this tag

driver.get(init_url)
count = 2
name_count = 2
page = 0

while page < 148:
    response = requests.get(init_url, headers = user_agent)
    soup = BeautifulSoup(response.text, 'html.parser')
    for movie in soup.find_all('a', class_ = 'title'):
        worksheet.write(get_column_letter(count) + '1', movie.text)
    
        print(movie.text, count)
        movie_arr.append(movie.text)
        url = "https://www.metacritic.com" + movie.get('href') + "/user-reviews"
        response = requests.get(url, headers = user_agent)
        soup = BeautifulSoup(response.text, 'html.parser')  
        driver.get(url)
        time.sleep(0.1)
    
        for review in soup.find_all('div', class_ = 'review pad_top1'):
            name1 = review.find('span', class_ = 'author').text
            rating = review.find('div', class_ = 'metascore_w user large movie mixed indiv')
            if(rating!=None):
                rating = review.find('div', class_ = 'metascore_w user large movie mixed indiv').text
            elif(rating==None):
                rating = review.find('div', class_ = 'metascore_w user large movie negative indiv')
                if(rating!=None):
                    rating = review.find('div', class_ = 'metascore_w user large movie negative indiv').text
                elif(rating==None):
                    rating = review.find('div', class_ = 'metascore_w user large movie positive indiv perfect')
                    if(rating!=None):
                        rating = review.find('div', class_ = 'metascore_w user large movie positive indiv perfect').text
                    elif(rating==None):
                        rating = review.find('div', class_ = 'metascore_w user large movie positive indiv').text


            if(name1 not in user_arr):
                user_arr.append(name1)
                worksheet.write('A' + str(name_count), name1)
                worksheet.write(get_column_letter(count) + str(name_count), rating)
                name_count += 1
            else:
                index_of_name = user_arr.index(name1) + 2
                worksheet.write(get_column_letter(count) + str(index_of_name), rating)
        count = count + 1
    page += 1
    init_url = "https://www.metacritic.com/browse/movies/score/metascore/all/filtered?sort=desc&page=" + str(page)
    driver.get(init_url)
    time.sleep(2)



driver.back()
driver.back()

workbook.close()



    

    
