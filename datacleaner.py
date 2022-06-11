# cleaning the data
 
# import openpyxl module
import openpyxl
import xlsxwriter
from openpyxl.utils.cell import get_column_letter
 
# new workbook to write to
workbook = xlsxwriter.Workbook('cleaneddata.xlsx')
worksheet = workbook.add_worksheet()

# reading from the excel file
path = "excel.xlsx"
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active

def copyRow(row):
    col = 1
    data = sheet_obj.cell(row = row, column = col)
    while data.value != None:
        worksheet.write(get_column_letter(col) + str(row), data.value)
        col += 1
        data = sheet_obj.cell(row = row, column = col)
    

# adding the movies into the excel file
count = 2
cell_obj = sheet_obj.cell(row = 1, column = 2)
while cell_obj.value != None:
    worksheet.write(get_column_letter(count) + '1', cell_obj.value)
    cell_obj = sheet_obj.cell(row = 1, column = count)
    count += 1
 
user_count = 2
contin = True
while contin:
    cell_obj = sheet_obj.cell(row = user_count, column = 1)
    if cell_obj.value == None:
        contin = False
    else:
        iter = 0
        for i in range(2, count):
            rating = sheet_obj.cell(row = user_count, column = i)
            if rating.value != None:
                iter += 1
        if iter > 1:
            worksheet.write(get_column_letter(user_count) + '1', cell_obj.value)
            copyRow(user_count)
        user_count += 1
        print(user_count)



workbook.close()